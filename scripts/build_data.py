import json
import math
import statistics
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SITE = ROOT / "web-dashboard"
OUT = SITE / "data.js"

DASHBOARD = ROOT / "outputs" / "upsc-prelims-2026" / "UPSC_Prelims_PYQ_Dashboard_2011_2026.xlsx"
QUESTIONS_2026 = ROOT / "outputs" / "upsc-prelims-2026" / "UPSC_Prelims_2026_GS_Set_A_Questions.xlsx"
PYQ_MASTER = ROOT / "Gyangram_PYQ_Master_File.xlsx"
REPORT_HTML = Path(r"C:\Users\radha\Downloads\UPSC_Prelims_2026_Analysis_Report.html")


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\r", "\n").strip()
    if text.lower() == "nan":
        return ""
    return " ".join(text.split())


def rows_from_sheet(path, sheet_name, header_row=1):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [clean(c) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    rows = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(clean(v) for v in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
    return rows


subject_compare = rows_from_sheet(DASHBOARD, "2026 vs History")
subject_trend = rows_from_sheet(DASHBOARD, "Subject Trend")
syllabus_map = rows_from_sheet(DASHBOARD, "2026 Syllabus Map")
heatmap = rows_from_sheet(DASHBOARD, "Syllabus Heatmap")
top_topics = rows_from_sheet(DASHBOARD, "Top Topics")
questions = rows_from_sheet(QUESTIONS_2026, "Questions")
pyq = rows_from_sheet(PYQ_MASTER, "PYQ Master", 3)

for row in subject_compare:
    row["2026 Share"] = round(float(row.get("2026 Share") or 0) * 100, 1)

for row in syllabus_map:
    row["No"] = int(row["No"])
    row["Historical Topic Count"] = int(row.get("Historical Topic Count") or 0)
    row["Mapping Score"] = float(row.get("Mapping Score") or 0)

for row in heatmap:
    for key in ["2011-2025 Count", "2026 Count", "Total 2011-2026"]:
        row[key] = int(row.get(key) or 0)

hist_subject_counts = Counter(clean(r.get("Subject")) for r in pyq if clean(r.get("Subject")))
year_counts = Counter(int(r.get("Year")) for r in pyq if r.get("Year"))
difficulty_counts = Counter(clean(r.get("Difficulty")) or "Unlabelled" for r in pyq)
confidence_counts = Counter(clean(r.get("Mapping Confidence")) for r in syllabus_map)
signal_counts_2026 = Counter(clean(r.get("Topic Signal")) or "Unsignalled" for r in syllabus_map)
answer_status_counts = Counter(clean(r.get("Answer Status")) for r in syllabus_map if clean(r.get("Answer Status")))
changed_answer_rows = [r for r in syllabus_map if clean(r.get("Answer Status")) == "Changed from tentative key"]
dropped_answer_rows = [r for r in syllabus_map if clean(r.get("Answer Status")) == "Dropped by UPSC"]

pyq_trace_rows = []
for row in pyq:
    year = int(row.get("Year") or 0)
    qnum = int(row.get("Q#") or 0)
    pyq_trace_rows.append({
        "Q#": qnum,
        "Year": year,
        "Set Question": qnum - 1500 if year == 2026 and qnum > 1500 else "",
        "Subject": clean(row.get("Subject")) or "Unmapped",
        "Unit": clean(row.get("Unit")) or "Unmapped",
        "Topic": clean(row.get("Topic")) or "Unmapped",
        "Topic ID": clean(row.get("Topic ID")) or "UNMAPPED",
        "Question": clean(row.get("Full Question")),
        "Option A": clean(row.get("Option A")),
        "Option B": clean(row.get("Option B")),
        "Option C": clean(row.get("Option C")),
        "Option D": clean(row.get("Option D")),
        "Answer": clean(row.get("Correct Answer")),
        "Explanation": clean(row.get("Explanation")),
        "Difficulty": clean(row.get("Difficulty")) or "Unlabelled",
    })

topic_summary = {}
for row in pyq_trace_rows:
    key = (row["Subject"], row["Unit"], row["Topic"], row["Topic ID"])
    item = topic_summary.setdefault(key, {
        "Subject": row["Subject"],
        "Unit": row["Unit"],
        "Topic": row["Topic"],
        "Topic ID": row["Topic ID"],
        "Total PYQs": 0,
        "PYQs 2011-2025": 0,
        "PYQs 2026": 0,
        "Years": set(),
    })
    item["Total PYQs"] += 1
    if row["Year"] == 2026:
        item["PYQs 2026"] += 1
    else:
        item["PYQs 2011-2025"] += 1
    item["Years"].add(row["Year"])

topic_summary_rows = []
for item in topic_summary.values():
    years = sorted(y for y in item.pop("Years") if y)
    item["Unique Years"] = len(years)
    item["Years Asked"] = ", ".join(str(y) for y in years)
    topic_summary_rows.append(item)
topic_summary_rows.sort(key=lambda r: (-r["Total PYQs"], r["Subject"], r["Unit"], r["Topic"]))

subject_tops = defaultdict(Counter)
for row in syllabus_map:
    subject_tops[clean(row.get("Subject"))][clean(row.get("Topic"))] += 1

topic_heat = defaultdict(lambda: {"subject": "", "unit": "", "topics": 0, "q2026": 0, "historical": 0})
for row in heatmap:
    key = (clean(row.get("Subject")), clean(row.get("Unit")))
    topic_heat[key]["subject"] = key[0]
    topic_heat[key]["unit"] = key[1]
    topic_heat[key]["topics"] += 1
    topic_heat[key]["q2026"] += row["2026 Count"]
    topic_heat[key]["historical"] += row["2011-2025 Count"]

rare_2026 = [
    row for row in syllabus_map
    if int(row.get("Historical Topic Count") or 0) <= 1 or "Not yet asked" in clean(row.get("Topic Signal"))
]

codex_findings = [
    {
        "label": "Pattern Shift",
        "title": "Science & Technology is the clearest 2026 over-index.",
        "detail": "It appears at 2.28x its 2011-2025 yearly average, making it the strongest subject-level signal in the 2026 paper.",
        "metric": "2.28x",
    },
    {
        "label": "Mastery Loop",
        "title": "The syllabus map turns every PYQ into an actionable topic node.",
        "detail": "Each 2026 question is linked to a unit, topic, topic ID, signal, answer, and explanation so GyanGram can move from question solving to diagnosis.",
        "metric": "100 mapped",
    },
    {
        "label": "Coverage Risk",
        "title": "Rare and previously untapped nodes matter in 2026.",
        "detail": "The dashboard flags low-history topic hits so content strategy does not become over-dependent on frequency alone.",
        "metric": f"{len(rare_2026)} flagged",
    },
    {
        "label": "Audit Layer",
        "title": "Algorithmic mapping is reviewable, not hidden.",
        "detail": "Confidence and score fields expose where human review should focus before production import into GyanGram.",
        "metric": f"{confidence_counts.get('Low', 0)} low confidence",
    },
]

report_insights = {
    "headline": "The Paper that Rewrote the Rules",
    "tldr": "UPSC 2026 was a Science & Technology paper wearing a History coat. It raised S&T, revived Art & Culture, and introduced ethics-style case scenarios into Prelims.",
    "sourceFile": str(REPORT_HTML),
    "compare2025": [
        {"subject": "Science & Technology", "y2025": 13, "y2026": 19, "delta": 6, "direction": "Sharp rise"},
        {"subject": "Art & Culture", "y2025": 2, "y2026": 9, "delta": 7, "direction": "Strong rebound"},
        {"subject": "International Relations", "y2025": 8, "y2026": 12, "delta": 4, "direction": "Up"},
        {"subject": "Social Issues / Schemes", "y2025": 3, "y2026": 4, "delta": 1, "direction": "Marginal"},
        {"subject": "Economy", "y2025": 18, "y2026": 16, "delta": -2, "direction": "Slight dip"},
        {"subject": "History (Ancient + Modern)", "y2025": 13, "y2026": 12, "delta": -1, "direction": "Stable"},
        {"subject": "Geography (all)", "y2025": 13, "y2026": 9, "delta": -4, "direction": "Down"},
        {"subject": "Indian Polity", "y2025": 14, "y2026": 9, "delta": -5, "direction": "Sharp drop"},
        {"subject": "Environment & Ecology", "y2025": 15, "y2026": 10, "delta": -5, "direction": "Sharp drop"},
    ],
    "questionTypes": [
        {"type": "Statement-based", "count": 45},
        {"type": "Direct / single-best answer", "count": 23},
        {"type": "Multi-select", "count": 22},
        {"type": "Match the following", "count": 4},
        {"type": "Pair matching", "count": 4},
        {"type": "How-many-of-the-above", "count": 2},
    ],
    "currentStatic": [
        {"subject": "Science & Tech", "current": 10, "static": 9},
        {"subject": "Environment", "current": 6, "static": 4},
        {"subject": "Economy", "current": 5, "static": 11},
        {"subject": "Social/Schemes", "current": 3, "static": 1},
        {"subject": "IR", "current": 3, "static": 9},
        {"subject": "Art & Culture", "current": 2, "static": 7},
        {"subject": "Polity", "current": 1, "static": 8},
        {"subject": "World Geo", "current": 2, "static": 1},
        {"subject": "Physical Geo", "current": 1, "static": 1},
        {"subject": "Modern Hist", "current": 1, "static": 4},
        {"subject": "Ancient Hist", "current": 0, "static": 7},
        {"subject": "Indian Geo", "current": 0, "static": 4},
    ],
    "shifts": [
        {
            "tag": "Shift 01 · S&T surge",
            "title": "The Atmanirbhar exam",
            "detail": "Nineteen S&T questions tilted toward indigenous capability: DHRUV64, GenomeIndia, National Quantum Mission, Bharat Forecast System, Deep Ocean Mission, Mission Sudarshan Chakra, Agnikul and Skyroot.",
        },
        {
            "tag": "Shift 02 · Culture rebound",
            "title": "Culture returns from the dead",
            "detail": "Art & Culture bounced from a very low 2025 base to nine questions in 2026. The lesson is not to zero out any subject after one quiet year.",
        },
        {
            "tag": "Shift 03 · Polity compression",
            "title": "Five fewer Polity questions",
            "detail": "Polity fell to nine questions but became trickier through conclusion-style option sets that punish pattern elimination.",
        },
        {
            "tag": "Shift 04 · Environment cools",
            "title": "Environment loses safety-net status",
            "detail": "Environment fell to ten questions, changing its study-hour return profile while still remaining strategically relevant.",
        },
    ],
    "mutations": [
        {
            "title": "Ethics scenarios enter Prelims",
            "detail": "Questions 51-53 read like GS-IV mini case studies around accountability, multi-stakeholder dialogue, limited disclosure and conflict resolution.",
        },
        {
            "title": "Conclusion-style options",
            "detail": "New option families such as 'There is no correct statement' force statement-by-statement verification.",
        },
        {
            "title": "How-many traps mature",
            "detail": "Dense pair/table questions require verifying both sides of a pair, not just recognizing one familiar name.",
        },
        {
            "title": "Questions are getting longer",
            "detail": "Average question length was estimated at 421 characters, with long Polity and Social Issues passages turning Prelims into a reading-speed test.",
        },
    ],
    "topicPatterns": [
        "Frontier technology with an Indian address: indigenous missions, platforms, chips, space and strategic technology.",
        "Economy as digital finance and policy design: UPI, CBDC, ONDC, TReDS, bonds, FI-Index and market plumbing.",
        "International Relations as connectivity and multilateralism: forums, corridors, UN bodies, neighbours and India-backed projects.",
        "The paper is moving away from plain general awareness toward conceptual plus contemporary integration.",
    ],
    "prepRecommendations": [
        "Treat Science & Technology as a core subject with a tracker for missions, indigenous platforms, private space and frontier technologies.",
        "Re-add Art & Culture to the weekly schedule: music, iconography, cave paintings, sculpture schools, folk traditions and heritage sites.",
        "Practice conclusion-style options by marking every statement true or false before looking at the answer codes.",
        "Prepare ethics-scenario Prelims questions using accountability, conflict of interest, mediation, disclosure and decision-making vocabulary.",
        "Read Economy through digital finance: RBI payments, SEBI updates, DPI, fintech, ONDC, CBDC and bond categories.",
        "Build a global-affairs project map for neighbours, corridors, multilateral forums, UN platforms and India-funded projects.",
        "Train for length with strict timed mocks; reading speed is now a UPSC skill.",
    ],
    "caveats": [
        "One year is one data point. The reliable inference is no subject zero, not that 2026 is the permanent pattern.",
        "Subject tags are heuristic because several 2026 questions cross syllabus boundaries.",
        "Answer keys and 2026 mappings remain tentative until official UPSC keys and manual topic review.",
        "Set A order differs from other sets, but subject and pattern findings hold across sets.",
    ],
}

payload = {
    "meta": {
        "title": "UPSC Prelims PYQ Intelligence Dashboard",
        "subtitle": "2011-2026 analysis linked to GyanGram syllabus mastery",
        "generatedFrom": [
            "Gyangram_PYQ_Master_File.xlsx",
            "Gyangram_Syllabus_Master_file.xlsx",
            "UPSC_Prelims_2026_GS_Set_A_Questions.xlsx",
            "UPSC provisional answer key dated 27-05-2026",
        ],
        "totals": {
            "historicalQuestions": len(pyq),
            "questions2026": len(syllabus_map),
            "totalPyqs": len(pyq),
            "subjects2026": len({clean(r.get("Subject")) for r in syllabus_map}),
            "topics2026": len({clean(r.get("Topic ID")) for r in syllabus_map}),
            "rare2026": len(rare_2026),
            "dropped2026": len(dropped_answer_rows),
        },
    },
    "answerKey": {
        "source": "UPSC provisional answer key, GS-I CSP Exam 2026, Series A, dated 27-05-2026",
        "changedCount": len(changed_answer_rows),
        "changedQuestions": [int(r.get("No")) for r in changed_answer_rows],
        "droppedCount": len(dropped_answer_rows),
        "droppedQuestions": [int(r.get("No")) for r in dropped_answer_rows],
        "statusCounts": answer_status_counts,
    },
    "subjectCompare": subject_compare,
    "subjectTrend": subject_trend,
    "syllabusMap": syllabus_map,
    "heatmap": heatmap,
    "topTopics": top_topics,
    "questions": questions,
    "pyqTraceRows": pyq_trace_rows,
    "topicSummaryRows": topic_summary_rows,
    "counts": {
        "historicalSubjects": hist_subject_counts,
        "yearCounts": year_counts,
        "difficulty": difficulty_counts,
        "confidence": confidence_counts,
        "topicSignals2026": signal_counts_2026,
    },
    "unitHeat": list(topic_heat.values()),
    "subjectTopTopics": {k: v.most_common(8) for k, v in subject_tops.items()},
    "rare2026": rare_2026,
    "codexFindings": codex_findings,
    "reportInsights": report_insights,
}


def sanitize(obj):
    if isinstance(obj, dict):
        return {str(k): sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [sanitize(v) for v in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    return obj


OUT.write_text(
    "window.DASHBOARD_DATA = "
    + json.dumps(sanitize(payload), ensure_ascii=False, indent=2)
    + ";\n",
    encoding="utf-8",
)
print(f"Wrote {OUT}")
